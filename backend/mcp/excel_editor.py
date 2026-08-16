"""
excel_editor.py — High-performance backend engine for the In-Browser Spreadsheet Editor.

Supports:
- Reading complete sheet grid data with headers, rows, sheet tabs, and dimensions
- Single cell & batch updates with atomic writes and Windows file-lock resilience
- Row insert, append, and delete operations
- Column header additions and renaming
- Worksheet tab creation and deletion
- New workbook creation with customizable starter schemas
"""

import logging
import os
import shutil
import tempfile
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from backend.mcp.excel_server import _resolve_path, _load_workbook, _get_headers


def get_sheet_data(
    workbook_path: str,
    worksheet_name: Optional[str] = None,
    page: int = 1,
    limit: int = 200,
) -> dict:
    """
    Read spreadsheet structure and data rows.
    Returns:
      sheet_names, active_sheet, headers, rows (1-indexed row numbers + values list),
      total_rows, total_cols, filename.
    """
    path = _resolve_path(workbook_path)
    wb = _load_workbook(str(path), read_only=True)
    
    sheet_names = wb.sheetnames
    target_sheet = worksheet_name if worksheet_name and worksheet_name in sheet_names else wb.sheetnames[0]
    ws = wb[target_sheet]

    # Collect headers from row 1
    raw_headers = []
    max_col = ws.max_column or 1
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=1, column=col_idx).value
        raw_headers.append(str(val) if val is not None else f"Column {get_column_letter(col_idx)}")

    # Trim trailing purely empty generated headers if any, but ensure at least 1 column
    while len(raw_headers) > 1 and raw_headers[-1].startswith("Column ") and ws.cell(row=1, column=len(raw_headers)).value is None:
        raw_headers.pop()

    headers = raw_headers if raw_headers else ["Column A"]
    col_count = len(headers)
    total_rows = ws.max_row or 1

    # Read rows (row 2 onwards for data)
    start_row = 2 + (page - 1) * limit
    end_row = min(total_rows + 1, start_row + limit)

    rows = []
    for r in range(2, total_rows + 1):
        row_vals = []
        is_empty = True
        for c in range(1, col_count + 1):
            cell_val = ws.cell(row=r, column=c).value
            if cell_val is not None and str(cell_val).strip() != "":
                is_empty = False
            row_vals.append("" if cell_val is None else str(cell_val))
        
        rows.append({
            "row_index": r,
            "values": row_vals,
            "is_empty": is_empty,
        })

    wb.close()

    return {
        "success": True,
        "workbook_path": str(path).replace("\\", "/"),
        "filename": path.name,
        "sheet_names": sheet_names,
        "active_sheet": target_sheet,
        "headers": headers,
        "rows": rows,
        "total_rows": max(len(rows) + 1, total_rows),
        "total_columns": col_count,
    }


def _save_workbook_atomic(wb: openpyxl.Workbook, target_path: Path):
    """Save an openpyxl Workbook atomically to avoid file lock issues."""
    tmp_write = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=target_path.parent)
    tmp_write.close()
    wb.save(tmp_write.name)
    wb.close()

    last_err = None
    import time
    for attempt in range(5):
        try:
            os.replace(tmp_write.name, target_path)
            return
        except PermissionError as e:
            last_err = e
            time.sleep(0.15)
    
    # Fallback to copy if os.replace is persistently held
    try:
        shutil.copy2(tmp_write.name, target_path)
        Path(tmp_write.name).unlink(missing_ok=True)
    except Exception:
        if Path(tmp_write.name).exists():
            Path(tmp_write.name).unlink(missing_ok=True)
        raise last_err or PermissionError(f"Could not overwrite {target_path}")


def update_cell(
    workbook_path: str,
    worksheet_name: Optional[str],
    row: int,
    col: int,
    value: Any,
) -> dict:
    """Update a single cell at (row, col) (1-indexed)."""
    path = _resolve_path(workbook_path)
    wb = _load_workbook(str(path), read_only=False)
    
    target_sheet = worksheet_name if worksheet_name and worksheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target_sheet]

    ws.cell(row=row, column=col).value = value
    _save_workbook_atomic(wb, path)

    return {
        "success": True,
        "row": row,
        "col": col,
        "value": value,
        "message": f"Updated cell at Row {row}, Col {col}.",
    }


def update_batch(
    workbook_path: str,
    worksheet_name: Optional[str],
    updates: list[dict],
) -> dict:
    """
    Update multiple cells in a single atomic write operation.
    updates: [{ "row": int, "col": int, "value": Any }, ...]
    """
    path = _resolve_path(workbook_path)
    wb = _load_workbook(str(path), read_only=False)

    target_sheet = worksheet_name if worksheet_name and worksheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target_sheet]

    for u in updates:
        r = int(u["row"])
        c = int(u["col"])
        v = u.get("value", "")
        ws.cell(row=r, column=c).value = v

    _save_workbook_atomic(wb, path)

    return {
        "success": True,
        "updated_count": len(updates),
        "message": f"Successfully updated {len(updates)} cells.",
    }


def add_row(
    workbook_path: str,
    worksheet_name: Optional[str] = None,
    row_data: Optional[list[Any]] = None,
    insert_at_row: Optional[int] = None,
) -> dict:
    """Append or insert a new row in the worksheet."""
    path = _resolve_path(workbook_path)
    wb = _load_workbook(str(path), read_only=False)

    target_sheet = worksheet_name if worksheet_name and worksheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target_sheet]

    if insert_at_row and insert_at_row <= ws.max_row:
        ws.insert_rows(insert_at_row)
        target_r = insert_at_row
    else:
        target_r = ws.max_row + 1

    if row_data:
        for c_idx, val in enumerate(row_data, start=1):
            ws.cell(row=target_r, column=c_idx).value = val

    _save_workbook_atomic(wb, path)

    return {
        "success": True,
        "row_index": target_r,
        "total_rows": ws.max_row,
        "message": f"Added new row at index {target_r}.",
    }


def delete_row(
    workbook_path: str,
    worksheet_name: Optional[str],
    row_index: int,
) -> dict:
    """Delete a row from the worksheet (1-indexed)."""
    if row_index < 2:
        raise ValueError("Cannot delete header row (Row 1).")

    path = _resolve_path(workbook_path)
    wb = _load_workbook(str(path), read_only=False)

    target_sheet = worksheet_name if worksheet_name and worksheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target_sheet]

    if row_index > ws.max_row:
        raise ValueError(f"Row {row_index} does not exist.")

    ws.delete_rows(row_index)
    _save_workbook_atomic(wb, path)

    return {
        "success": True,
        "deleted_row": row_index,
        "remaining_rows": ws.max_row,
        "message": f"Deleted row {row_index}.",
    }


def add_column(
    workbook_path: str,
    worksheet_name: Optional[str],
    column_name: str,
) -> dict:
    """Append a new column with the given header to row 1."""
    if not column_name or not column_name.strip():
        raise ValueError("Column name cannot be empty.")

    path = _resolve_path(workbook_path)
    wb = _load_workbook(str(path), read_only=False)

    target_sheet = worksheet_name if worksheet_name and worksheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target_sheet]

    headers = _get_headers(ws)
    new_col_idx = len(headers) + 1
    ws.cell(row=1, column=new_col_idx).value = column_name.strip()

    _save_workbook_atomic(wb, path)

    return {
        "success": True,
        "column_name": column_name.strip(),
        "column_index": new_col_idx,
        "message": f"Added column '{column_name.strip()}'.",
    }


def create_sheet(
    workbook_path: str,
    sheet_name: str,
) -> dict:
    """Create a new worksheet tab in the workbook."""
    if not sheet_name or not sheet_name.strip():
        raise ValueError("Sheet name cannot be empty.")

    path = _resolve_path(workbook_path)
    wb = _load_workbook(str(path), read_only=False)

    clean_name = sheet_name.strip()
    if clean_name in wb.sheetnames:
        raise ValueError(f"Sheet '{clean_name}' already exists.")

    ws = wb.create_sheet(title=clean_name)
    # Give default header
    ws.cell(row=1, column=1).value = "Item"
    ws.cell(row=1, column=2).value = "Details"

    _save_workbook_atomic(wb, path)

    return {
        "success": True,
        "sheet_name": clean_name,
        "sheet_names": wb.sheetnames,
        "message": f"Created worksheet tab '{clean_name}'.",
    }


def create_new_workbook(
    filename: str,
    headers: Optional[list[str]] = None,
    sheet_name: str = "Sheet1",
) -> dict:
    """Create a brand new .xlsx workbook in the uploads/ directory."""
    if not filename:
        filename = "Untitled_Spreadsheet.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    root_dir = Path(__file__).parent.parent.parent
    uploads_dir = root_dir / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)

    file_path = uploads_dir / filename
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name or "Sheet1"

    default_headers = headers or ["Name", "Description", "Category", "Value", "Status", "Notes"]
    for col_idx, h in enumerate(default_headers, start=1):
        ws.cell(row=1, column=col_idx).value = h

    # Add an initial empty row template
    for col_idx in range(1, len(default_headers) + 1):
        ws.cell(row=2, column=col_idx).value = ""

    _save_workbook_atomic(wb, file_path)

    return {
        "success": True,
        "filename": filename,
        "workbook_path": f"./uploads/{filename}",
        "sheet_name": ws.title,
        "headers": default_headers,
        "message": f"Created new workbook '{filename}'.",
    }
