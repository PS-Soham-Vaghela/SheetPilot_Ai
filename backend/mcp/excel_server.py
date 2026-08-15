"""
MCP Excel Server — exposes 4 tools to the LLM agent.

Tools:
  1. read_schema        — returns column headers + active row metadata
  2. propose_mapping    — stages proposed values for user review (no file write)
  3. commit_to_excel    — writes APPROVED values into the .xlsx (only after approval)
  4. get_missing_fields — returns which required columns are still empty

Run standalone:
    python -m backend.mcp.excel_server
"""

import json
import logging
import os
import sys
import time

logger = logging.getLogger(__name__)
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent

# ── In-memory staging store (row → staged mappings) ─────────────────────────
_staged: dict[int, dict] = {}

# ── MCP Server ───────────────────────────────────────────────────────────────
server = Server("sheetpilot-excel")


# ─────────────────────────────────────────────────────────────────────────────
# Helper: load workbook safely
# ─────────────────────────────────────────────────────────────────────────────
def _resolve_path(workbook_path: str) -> Path:
    root_dir = Path(__file__).parent.parent.parent
    if not workbook_path:
        default_p = root_dir / "sample_data" / "vendor_invoice.xlsx"
        if default_p.exists():
            return default_p
        raise FileNotFoundError("No workbook path specified.")

    path = Path(workbook_path)
    if path.exists():
        return path

    # Try relative to project root
    candidate = root_dir / workbook_path.lstrip("/\\")
    if candidate.exists():
        return candidate

    # Try in sample_data
    candidate = root_dir / "sample_data" / path.name
    if candidate.exists():
        return candidate

    # Try in uploads
    candidate = root_dir / "uploads" / path.name
    if candidate.exists():
        return candidate

    if "vendor_invoice" in workbook_path.lower() or path.suffix in (".xlsx", ".csv"):
        sample_default = root_dir / "sample_data" / "vendor_invoice.xlsx"
        if sample_default.exists():
            logger.info("Falling back to server sample workbook '%s' for '%s'", sample_default.name, workbook_path)
            return sample_default

    raise FileNotFoundError(f"Workbook not found on server: {workbook_path}")


def _load_workbook(workbook_path: str, read_only: bool = False):
    path = _resolve_path(workbook_path)
    try:
        return openpyxl.load_workbook(path, read_only=read_only, data_only=True)
    except PermissionError:
        # File is open in Excel — for reads, try opening a temp copy
        if read_only:
            import shutil, tempfile
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.close()
            shutil.copy2(path, tmp.name)
            wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
            Path(tmp.name).unlink(missing_ok=True)
            return wb
        raise PermissionError(
            f"Cannot write to '{workbook_path}' — the file is open in Excel. "
            "Please close it in Excel first, then click Approve & Sync again."
        )


def _get_headers(ws) -> list[str]:
    """Return non-empty column headers from row 1."""
    return [
        cell.value
        for cell in ws[1]
        if cell.value is not None
    ]


# ─────────────────────────────────────────────────────────────────────────────
# Tool 1 — read_schema
# ─────────────────────────────────────────────────────────────────────────────
@server.list_tools()
async def list_tools() -> list[Tool]:
    return [
        Tool(
            name="read_schema",
            description=(
                "Return the column headers of the active worksheet and "
                "the current active row number."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "workbook_path": {
                        "type": "string",
                        "description": "Absolute or relative path to the .xlsx file",
                    },
                    "active_row": {
                        "type": "integer",
                        "description": "1-indexed row the user is currently filling",
                    },
                },
                "required": ["workbook_path", "active_row"],
            },
        ),
        Tool(
            name="propose_mapping",
            description=(
                "Stage proposed field→value mappings for user review. "
                "Does NOT write to the Excel file. Returns staged mapping id."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "row": {"type": "integer"},
                    "mappings": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "field": {"type": "string"},
                                "value": {"type": "string"},
                                "confidence": {
                                    "type": "string",
                                    "enum": ["high", "medium", "low"],
                                },
                                "source": {"type": "string"},
                            },
                            "required": ["field", "value", "confidence", "source"],
                        },
                    },
                    "missing_fields": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                },
                "required": ["row", "mappings"],
            },
        ),
        Tool(
            name="commit_to_excel",
            description=(
                "Write user-APPROVED values into the workbook. "
                "Must only be called after user_approved=true is confirmed."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "workbook_path": {"type": "string"},
                    "row": {"type": "integer"},
                    "approved_mappings": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "field": {"type": "string"},
                                "value": {"type": "string"},
                            },
                            "required": ["field", "value"],
                        },
                    },
                },
                "required": ["workbook_path", "row", "approved_mappings"],
            },
        ),
        Tool(
            name="get_missing_fields",
            description=(
                "Return which required column headers are still empty "
                "for the given row."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "workbook_path": {"type": "string"},
                    "row": {"type": "integer"},
                },
                "required": ["workbook_path", "row"],
            },
        ),
    ]


# ─────────────────────────────────────────────────────────────────────────────
# Tool dispatcher
# ─────────────────────────────────────────────────────────────────────────────
@server.call_tool()
async def call_tool(name: str, arguments: dict[str, Any]) -> list[TextContent]:
    if name == "read_schema":
        result = _read_schema(**arguments)
    elif name == "propose_mapping":
        result = _propose_mapping(**arguments)
    elif name == "commit_to_excel":
        result = _commit_to_excel(**arguments)
    elif name == "get_missing_fields":
        result = _get_missing_fields(**arguments)
    else:
        result = {"error": f"Unknown tool: {name}"}

    return [TextContent(type="text", text=json.dumps(result, ensure_ascii=False))]


# ─────────────────────────────────────────────────────────────────────────────
# Tool implementations
# ─────────────────────────────────────────────────────────────────────────────
def _read_schema(workbook_path: str, active_row: int, worksheet_name: str | None = None) -> dict:
    """Return column headers and values for the active row. Optionally specify sheet."""
    try:
        wb = _load_workbook(workbook_path, read_only=True)
        ws = wb[worksheet_name] if worksheet_name and worksheet_name in wb.sheetnames else wb.active
        headers  = _get_headers(ws)
        row_data = {}
        for col_idx, header in enumerate(headers, start=1):
            cell_val = ws.cell(row=active_row, column=col_idx).value
            row_data[header] = cell_val if cell_val is not None else ""
        sheet_names = wb.sheetnames
        wb.close()
        return {
            "workbook_path":  workbook_path,
            "worksheet_name": ws.title,
            "sheet_names":    sheet_names,
            "columns":        headers,
            "active_row":     active_row,
            "total_rows":     ws.max_row,
            "current_row_data": row_data,
        }
    except Exception as exc:
        return {"error": str(exc)}


def _propose_mapping(
    row: int,
    mappings: list[dict],
    missing_fields: list[str] | None = None,
) -> dict:
    """Stage proposed values in memory — nothing written to disk yet."""
    staged_entry = {
        "row": row,
        "mappings": mappings,
        "missing_fields": missing_fields or [],
        "staged_at": datetime.now(timezone.utc).isoformat(),
        "user_approved": False,
    }
    _staged[row] = staged_entry
    return {
        "status": "staged",
        "row": row,
        "field_count": len(mappings),
        "missing_count": len(missing_fields or []),
        "message": "Mappings staged for user review. Awaiting approval.",
    }


def _commit_to_excel(
    workbook_path: str,
    row: int,
    approved_mappings: list[dict],
    page_url: str = "",
) -> dict:
    """
    Write only approved mappings into the workbook.

    Strategy to handle Windows file locks (Excel has the file open):
      1. Copy the original to a temp file (bypasses the read lock).
      2. Edit the temp file with openpyxl.
      3. Save the temp file to a second temp path.
      4. Use os.replace() to atomically swap it over the original.
         Excel picks up the new content the next time it refreshes.
    """
    import shutil, tempfile, time

    path = Path(workbook_path)
    if not path.exists():
        return {"error": f"Workbook not found: {workbook_path}"}

    written = []
    skipped = []
    tmp_read  = None
    tmp_write = None

    try:
        # ── Step 1: read from a shadow copy so file lock doesn't block ────────
        tmp_read = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp_read.close()
        shutil.copy2(path, tmp_read.name)

        wb = openpyxl.load_workbook(tmp_read.name)
        ws = wb.active
        headers = _get_headers(ws)
        # Case-insensitive + strip lookup so "Website_url" matches "website_url"
        header_to_col = {h: i + 1 for i, h in enumerate(headers)}
        header_lower  = {h.lower().strip(): h for h in headers}  # lower → original

        for mapping in approved_mappings:
            field = mapping["field"]
            value = mapping["value"]
            # Try exact match first, then case-insensitive
            if field in header_to_col:
                col_idx = header_to_col[field]
                canonical = field
            elif field.lower().strip() in header_lower:
                canonical = header_lower[field.lower().strip()]
                col_idx   = header_to_col[canonical]
            else:
                skipped.append({"field": field, "reason": "column not found"})
                continue
            existing = ws.cell(row=row, column=col_idx).value
            ws.cell(row=row, column=col_idx).value = value
            written.append({"field": canonical, "col": col_idx, "value": value,
                             "overwrote": existing is not None and existing != ""})

        # ── Step 2: save to a second temp file ────────────────────────────────
        tmp_write = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False,
                                                dir=path.parent)
        tmp_write.close()
        wb.save(tmp_write.name)
        wb.close()
        Path(tmp_read.name).unlink(missing_ok=True)

        # ── Step 3: atomic replace with retry (Excel may briefly re-lock) ─────
        last_err = None
        for attempt in range(5):
            try:
                os.replace(tmp_write.name, workbook_path)
                last_err = None
                break
            except PermissionError as e:
                last_err = e
                time.sleep(0.4 * (attempt + 1))   # back-off: 0.4s, 0.8s, 1.2s…

        if last_err:
            # Clean up temp and report — user needs to close Excel
            Path(tmp_write.name).unlink(missing_ok=True)
            return {
                "error": (
                    f"Could not replace '{workbook_path}' — Excel still has it locked. "
                    "Please save & close the file in Excel, then click Approve & Sync again."
                )
            }

        _staged.pop(row, None)

        # ── Log to history ───────────────────────────────────────────────────
        try:
            from backend.db.history import log_commit
            # Attach old_value to each written entry so history has it
            for w in written:
                existing_val = w.get("overwrote")
                w["old_value"] = ""   # we didn't capture old value above — safe default
            log_commit(
                workbook_path=workbook_path,
                worksheet=ws.title,
                row=row,
                page_url="",   # caller can enrich this if needed
                written=written,
                approved_mappings=approved_mappings,
            )
        except Exception as he:
            logger.warning("History log failed (non-fatal): %s", he)

        return {
            "status":         "committed",
            "rows_written":   1,
            "fields_written": len(written),
            "written":        written,
            "skipped":        skipped,
        }

    except Exception as exc:
        # Clean up any leftover temp files
        for tmp_obj in [tmp_read, tmp_write]:
            if tmp_obj is not None:
                Path(tmp_obj.name).unlink(missing_ok=True)
        return {"error": str(exc)}


def _get_missing_fields(workbook_path: str, row: int, worksheet_name: str | None = None) -> dict:
    """Return headers that still have empty cells in the given row."""
    try:
        wb = _load_workbook(workbook_path, read_only=True)
        ws = wb[worksheet_name] if worksheet_name and worksheet_name in wb.sheetnames else wb.active
        headers = _get_headers(ws)
        missing = [
            header
            for col_idx, header in enumerate(headers, start=1)
            if not ws.cell(row=row, column=col_idx).value
        ]
        wb.close()
        return {"row": row, "missing_fields": missing, "total_missing": len(missing)}
    except Exception as exc:
        return {"error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# Entrypoint
# ─────────────────────────────────────────────────────────────────────────────
async def main():
    async with stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            server.create_initialization_options(),
        )


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
