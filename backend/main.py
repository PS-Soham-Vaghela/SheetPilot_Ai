"""
main.py — FastAPI backend for SheetPilot AI.

Endpoints (original):
  POST /analyze          — RAG+LLM, returns proposals
  GET  /analyze/stream   — SSE streaming progress
  POST /commit           — write approved mappings to Excel/CSV
  GET  /schema           — Excel column headers + sheet list
  GET  /health           — liveness check
  GET  /history          — sync history
  POST /undo             — revert last commit
  GET  /suggest-mappings — dynamic semantic map
  GET  /next-empty-row   — returns next row with empty cells

Endpoints (new — web dashboard):
  POST /auth/register    — create account
  POST /auth/login       — get JWT token
  GET  /me               — current user profile
  GET  /dashboard/stats  — aggregate counts for overview cards
  GET  /dashboard/activity — syncs-per-day for chart
  GET  /dashboard/recent — last 10 syncs
  GET  /workbooks        — per-workbook summary
  POST /analyze-url      — server-side URL fetch + AI analyze
  GET  /history/search   — advanced filtered history
"""

import asyncio as _asyncio
import json
import logging
import os
import sys
from concurrent.futures import ThreadPoolExecutor as _TPE
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

load_dotenv(Path(__file__).parent.parent / ".env")

from backend.agent.orchestrator import handle_page_load, handle_user_approval
from backend.auth import (
    init_users_collection, register_user, login_user, get_current_user, get_user_by_id
)
from backend.db.history import (
    init_db, get_history, get_last_commit, mark_undone,
    get_stats, get_activity_by_day, get_workbooks_summary,
    get_recent_syncs, search_history,
)
from backend.llm.local_model import check_ollama_connection
from backend.mcp.agent_client import AgentMCPClient
from backend.mcp.excel_server import _commit_to_excel, _resolve_path
from backend.models.schemas import (
    PageContentRequest, ProposalResponse,
    CommitRequest, CommitResponse, ApprovedMapping,
)

_executor = _TPE(max_workers=4)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

# ── Webapp dist path ───────────────────────────────────────────────────────────
_WEBAPP_DIST = Path(__file__).parent.parent / "webapp" / "dist"


@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("SheetPilot AI backend starting up...")
    init_db()
    init_users_collection()
    logger.info("MongoDB collections + indexes initialised.")

    status = check_ollama_connection()
    if status["ok"]:
        logger.info("Groq API: %s", status["message"])
    else:
        logger.warning("Groq API: %s", status["message"])
    yield
    from backend.db.mongodb import close_db
    close_db()
    logger.info("SheetPilot AI shutting down.")


app = FastAPI(title="SheetPilot AI", version="3.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
_mcp = AgentMCPClient()


# ── Root: serve webapp or API info ────────────────────────────────────────────
@app.get("/", include_in_schema=False)
async def root():
    index = _WEBAPP_DIST / "index.html"
    if index.exists():
        return FileResponse(str(index))
    return {"service": "SheetPilot AI Backend v3", "docs": "/docs", "webapp": "not built yet"}


# ══════════════════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════════════════════════

class AuthRequest(BaseModel):
    email: str
    password: str


@app.post("/auth/register", tags=["Auth"])
async def register(body: AuthRequest):
    """Create a new account. Returns JWT token."""
    return register_user(body.email, body.password)


@app.post("/auth/login", tags=["Auth"])
async def login(body: AuthRequest):
    """Login with email + password. Returns JWT token."""
    return login_user(body.email, body.password)


@app.get("/me", tags=["Auth"])
async def me(user: dict = Depends(get_current_user)):
    """Return the currently authenticated user's profile."""
    return user


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/dashboard/stats", tags=["Dashboard"])
async def dashboard_stats():
    return get_stats()


@app.get("/dashboard/activity", tags=["Dashboard"])
async def dashboard_activity(days: int = 30):
    return {"activity": get_activity_by_day(days)}


@app.get("/dashboard/recent", tags=["Dashboard"])
async def dashboard_recent(limit: int = 10):
    return {"recent": get_recent_syncs(limit)}


# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOKS ROUTE
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/workbooks", tags=["Workbooks"])
@app.get("/api/workbooks", tags=["Workbooks"])
async def workbooks(request: Request):
    if "text/html" in request.headers.get("accept", "") and not request.url.path.startswith("/api/"):
        index = _WEBAPP_DIST / "index.html"
        if index.exists():
            return FileResponse(str(index))
    return {"workbooks": get_workbooks_summary()}


# ══════════════════════════════════════════════════════════════════════════════
# HISTORY ROUTES (extended)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/history", tags=["History"])
@app.get("/api/history", tags=["History"])
async def history(request: Request, workbook_path: Optional[str] = None, limit: int = 50):
    if "text/html" in request.headers.get("accept", "") and not request.url.path.startswith("/api/"):
        index = _WEBAPP_DIST / "index.html"
        if index.exists():
            return FileResponse(str(index))
    return {"history": get_history(workbook_path, limit)}


@app.get("/history/search", tags=["History"])
@app.get("/api/history/search", tags=["History"])
async def history_search(
    request: Request,
    q: str = "",
    workbook_path: str = "",
    page_url: str = "",
    date_from: str = "",
    date_to: str = "",
    limit: int = 25,
    offset: int = 0,
):
    if "text/html" in request.headers.get("accept", "") and not request.url.path.startswith("/api/"):
        index = _WEBAPP_DIST / "index.html"
        if index.exists():
            return FileResponse(str(index))
    return search_history(
        query=q,
        workbook_path=workbook_path,
        page_url=page_url,
        date_from=date_from,
        date_to=date_to,
        limit=limit,
        offset=offset,
    )


# ══════════════════════════════════════════════════════════════════════════════
# SERVER-SIDE URL ANALYZE
# ══════════════════════════════════════════════════════════════════════════════

class AnalyzeURLRequest(BaseModel):
    url: str
    workbook_path: str
    active_row: int = 2
    worksheet_name: Optional[str] = None


@app.post("/analyze-url", tags=["Agent"])
async def analyze_url(request: AnalyzeURLRequest):
    """
    Server-side URL fetch + AI analysis.
    Fetches the URL with httpx, extracts text with BeautifulSoup,
    then runs the full RAG+LLM pipeline — no browser extension needed.
    """
    import httpx
    from bs4 import BeautifulSoup

    try:
        async with httpx.AsyncClient(
            timeout=15.0,
            follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 SheetPilotBot/3.0"},
        ) as client:
            resp = await client.get(request.url)
            resp.raise_for_status()
    except httpx.HTTPError as e:
        raise HTTPException(status_code=422, detail=f"Failed to fetch URL: {e}")

    # Extract text from HTML with BeautifulSoup (mirrors extractText() in content_script.js)
    soup = BeautifulSoup(resp.text, "lxml")
    for tag in soup(["script", "style", "noscript", "svg", "canvas", "template"]):
        tag.decompose()

    meta_parts = []
    if soup.title:
        meta_parts.append("Page Title: " + soup.title.get_text().strip())
    canonical = soup.find("link", rel="canonical")
    if canonical:
        meta_parts.append("Canonical URL: " + canonical.get("href", ""))
    for m in soup.find_all("meta"):
        prop = m.get("property", "") or m.get("name", "")
        content = m.get("content", "")
        if prop and content:
            meta_parts.append(f"Meta {prop}: {content}")
    for h in soup.find_all(["h1", "h2"]):
        t = h.get_text(strip=True)
        if t:
            meta_parts.append(f"{h.name.upper()}: {t}")

    body_text = soup.get_text(separator=" ", strip=True)
    page_text = "\n".join(meta_parts) + "\n\n--- PAGE BODY ---\n" + body_text[:8000]

    try:
        loop = _asyncio.get_event_loop()
        staged = await loop.run_in_executor(
            _executor,
            lambda: handle_page_load(
                page_text=page_text,
                page_url=request.url,
                workbook_path=request.workbook_path,
                active_row=request.active_row,
                worksheet_name=request.worksheet_name,
            ),
        )
        return ProposalResponse(success=True, staged_mapping=staged)
    except Exception as exc:
        logger.exception("analyze-url pipeline error")
        raise HTTPException(status_code=500, detail=str(exc))


# ══════════════════════════════════════════════════════════════════════════════
# ORIGINAL EXTENSION ROUTES (unchanged)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/health", tags=["System"])
async def health():
    return {"status": "ok", "ollama": check_ollama_connection()}


@app.get("/schema", tags=["Excel"])
async def get_schema(workbook_path: str, active_row: int = 2,
                     worksheet_name: Optional[str] = None):
    result = _mcp.read_schema(workbook_path=workbook_path, active_row=active_row,
                               worksheet_name=worksheet_name)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@app.get("/next-empty-row", tags=["Excel"])
async def next_empty_row(workbook_path: str, start_row: int = 2,
                         worksheet_name: Optional[str] = None):
    try:
        from backend.mcp.excel_server import _load_workbook, _get_headers
        wb = _load_workbook(workbook_path, read_only=True)
        ws = wb[worksheet_name] if worksheet_name and worksheet_name in wb.sheetnames \
             else wb.active
        headers = _get_headers(ws)
        total = ws.max_row
        for r in range(start_row, total + 2):
            vals = [ws.cell(row=r, column=c+1).value for c in range(len(headers))]
            if any(v is None or str(v).strip() == "" for v in vals):
                wb.close()
                return {"next_empty_row": r, "total_rows": total}
        wb.close()
        return {"next_empty_row": total + 1, "total_rows": total}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/suggest-mappings", tags=["Agent"])
async def suggest_mappings(workbook_path: str, worksheet_name: Optional[str] = None):
    schema = _mcp.read_schema(workbook_path=workbook_path, active_row=2,
                               worksheet_name=worksheet_name)
    if "error" in schema:
        raise HTTPException(status_code=400, detail=schema["error"])
    columns = schema.get("columns", [])
    return {"columns": columns, "semantic_map": _build_semantic_map(columns)}


_HEURISTIC_PATTERNS = {
    "name": "company person name brand title",
    "url": "website URL https www link homepage",
    "website": "website URL https www link homepage",
    "email": "email contact address @",
    "phone": "phone number telephone mobile contact",
    "address": "address street city location postal",
    "description": "description about summary overview mission",
    "theme": "theme topic category industry focus",
    "price": "price cost fee amount currency",
    "date": "date year month time",
    "title": "title heading name role",
    "company": "company organisation business brand",
    "country": "country nation region location",
    "city": "city town location region",
    "linkedin": "linkedin social profile",
    "twitter": "twitter social handle",
    "industry": "industry sector category vertical",
}


def _build_semantic_map(columns: list[str]) -> dict[str, str]:
    result = {}
    for col in columns:
        key = col.lower().replace(" ", "_").replace("-", "_")
        if key in _HEURISTIC_PATTERNS:
            result[col] = _HEURISTIC_PATTERNS[key]
            continue
        matched = next(
            (v for k, v in _HEURISTIC_PATTERNS.items() if k in key or key in k),
            col.replace("_", " ").replace("-", " ")
        )
        result[col] = matched
    return result


@app.post("/analyze", response_model=ProposalResponse, tags=["Agent"])
async def analyze(request: PageContentRequest):
    logger.info("POST /analyze url=%s row=%d", request.page_url, request.active_row)
    try:
        loop = _asyncio.get_event_loop()
        staged = await loop.run_in_executor(
            _executor,
            lambda: handle_page_load(
                page_text=request.page_text,
                page_url=request.page_url,
                workbook_path=request.workbook_path,
                active_row=request.active_row,
                worksheet_name=getattr(request, "worksheet_name", None),
            )
        )
        return ProposalResponse(success=True, staged_mapping=staged)
    except RuntimeError as exc:
        return ProposalResponse(success=False, error=str(exc))
    except Exception as exc:
        logger.exception("Unexpected error in /analyze")
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/analyze/stream", tags=["Agent"])
async def analyze_stream(
    page_url: str, workbook_path: str, active_row: int,
    worksheet_name: Optional[str] = None,
):
    async def _event_gen():
        q: _asyncio.Queue = _asyncio.Queue()

        def _run():
            try:
                q.put_nowait({"type": "status", "msg": "Indexing page with RAG…"})
                from backend.mcp.excel_server import _read_schema, _get_missing_fields
                _read_schema(workbook_path, active_row, worksheet_name)
                q.put_nowait({"type": "status", "msg": "Schema loaded. Searching passages…"})
                missing_r = _get_missing_fields(workbook_path, active_row, worksheet_name)
                missing = missing_r.get("missing_fields", [])
                q.put_nowait({"type": "status", "msg": f"Found {len(missing)} empty fields. Asking AI…"})
                q.put_nowait({"type": "status", "msg": "AI reasoning in progress…"})
                q.put_nowait({"type": "done"})
            except Exception as exc:
                q.put_nowait({"type": "error", "msg": str(exc)})

        _executor.submit(_run)
        while True:
            evt = await q.get()
            yield "data: " + json.dumps(evt) + "\n\n"
            if evt["type"] in ("done", "error"):
                break

    return StreamingResponse(
        _event_gen(), media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


class ChatWorkbookRequest(BaseModel):
    workbook_path: str
    query: str


@app.post("/upload-workbook", tags=["Workbooks"])
async def upload_workbook(file: UploadFile = File(...)):
    """Upload an Excel (.xlsx) or CSV (.csv) workbook to the server."""
    uploads_dir = Path(__file__).parent.parent / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    file_path = uploads_dir / file.filename
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    return {
        "success": True,
        "filename": file.filename,
        "workbook_path": f"./uploads/{file.filename}",
        "message": f"Uploaded {file.filename} successfully.",
    }


@app.post("/chat/workbook", tags=["Chat"])
async def chat_workbook(request: ChatWorkbookRequest):
    logger.info("POST /chat/workbook query=%s workbook=%s", request.query, request.workbook_path)
    try:
        import openpyxl
        from backend.rag.retriever import index_page, search
        from backend.llm.local_model import chat
        
        try:
            path = _resolve_path(request.workbook_path)
        except Exception as exc:
            raise HTTPException(status_code=404, detail=str(exc))
            
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        row_texts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value is not None]
            if not headers:
                continue
            # Read first 100 rows to keep indexing fast
            for r_idx in range(2, min(ws.max_row + 1, 102)):
                row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, len(headers) + 1)]
                if not any(v is not None for v in row_vals):
                    continue
                row_desc = f"Sheet: {sheet_name}, Row {r_idx}: "
                parts = []
                for h, v in zip(headers, row_vals):
                    if v is not None:
                        parts.append(f"{h}={v}")
                if parts:
                    row_desc += ", ".join(parts)
                    row_texts.append(row_desc)
                    
        if not row_texts:
            return {"response": "The workbook appears to be empty or has no data rows."}
            
        workbook_text = "\n\n".join(row_texts)
        if len(workbook_text) < 8000:
            prompt = (
                f"The user is asking a question about their Excel workbook '{path.name}'.\n"
                f"Here is the complete spreadsheet data:\n\n{workbook_text}\n\n"
                f"User Question: {request.query}\n"
                f"Please answer the user's question accurately and concisely using only the spreadsheet data provided above."
            )
        else:
            index_page(page_text=workbook_text, page_url=str(path))
            passages = search(query=request.query, page_url=str(path), k=10)
            prompt = (
                f"The user is asking a question about their Excel workbook '{path.name}'.\n"
                f"Here are the most relevant rows retrieved from the spreadsheet:\n\n"
            )
            for p in passages:
                prompt += f"- {p.text}\n"
            prompt += (
                f"\nUser Question: {request.query}\n"
                f"Please answer the user's question accurately and concisely using only the spreadsheet data provided above."
            )
        
        llm_response = chat(prompt)
        return {"response": llm_response}
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Unexpected error in /chat/workbook")
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/commit", response_model=CommitResponse, tags=["Agent"])
async def commit(request: CommitRequest, workbook_path: str,
                 worksheet_name: Optional[str] = None,
                 page_url: Optional[str] = ""):
    logger.info("POST /commit workbook=%s row=%d", workbook_path, request.row)
    try:
        result = handle_user_approval(
            workbook_path=workbook_path, row=request.row,
            approved_mappings=request.approved_mappings, page_url=page_url or "",
        )
        return result
    except Exception as exc:
        logger.exception("Unexpected error in /commit")
        raise HTTPException(status_code=500, detail=str(exc))


class UndoRequest(BaseModel):
    workbook_path: str
    history_id: Optional[int] = None


@app.post("/undo", tags=["History"])
async def undo(request: UndoRequest):
    if request.history_id:
        records = [r for r in get_history(request.workbook_path, 200)
                   if r["id"] == request.history_id]
    else:
        last = get_last_commit(request.workbook_path)
        records = [last] if last else []

    if not records or records[0] is None:
        raise HTTPException(status_code=404, detail="No commit found to undo.")

    record = records[0]
    fields_json = json.loads(record["fields_json"])
    restore = [
        {"field": f["field"], "value": f["old_value"]}
        for f in fields_json if f.get("old_value", "") != ""
    ]
    if not restore:
        restore = [{"field": f["field"], "value": ""} for f in fields_json]

    result = _commit_to_excel(
        workbook_path=request.workbook_path,
        row=record["row"],
        approved_mappings=restore,
    )
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])

    mark_undone(record["id"])
    return {
        "success": True, "undone_row": record["row"],
        "fields": len(restore),
        "message": f"Row {record['row']} restored to previous values.",
    }


# ── Global exception handler ──────────────────────────────────────────────────
@app.exception_handler(Exception)
async def _global(request: Request, exc: Exception):
    logger.error("Unhandled: %s", exc)
    return JSONResponse(status_code=500, content={"error": str(exc)})


# ── Serve built webapp (must be AFTER all API routes) ─────────────────────────
if _WEBAPP_DIST.exists():
    app.mount("/assets", StaticFiles(directory=str(_WEBAPP_DIST / "assets")), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def spa_fallback(full_path: str):
        """Catch-all for SPA client-side routing."""
        index = _WEBAPP_DIST / "index.html"
        if index.exists():
            return FileResponse(str(index))
        raise HTTPException(status_code=404)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "backend.main:app",
        host=os.getenv("BACKEND_HOST", "127.0.0.1"),
        port=int(os.getenv("BACKEND_PORT", "8000")),
        reload=True,
    )
