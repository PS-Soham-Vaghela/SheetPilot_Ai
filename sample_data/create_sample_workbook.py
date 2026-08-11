"""
create_sample_workbook.py — generates vendor_invoice.xlsx for testing.

Run once:
    python sample_data/create_sample_workbook.py

The workbook has one sheet "Invoices" with realistic vendor/invoice headers.
Row 1 = headers. Rows 2+ are data rows (row 2 is the default active_row).
Some rows are intentionally partially filled so get_missing_fields has something to return.
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Column schema ─────────────────────────────────────────────────────────────
HEADERS = [
    "Vendor Name",
    "Vendor GST Number",
    "Invoice Number",
    "Invoice Date",
    "Due Date",
    "Item Description",
    "Quantity",
    "Unit Price",
    "Total Amount",
    "Currency",
    "Payment Terms",
    "Vendor Email",
    "Vendor Phone",
    "Vendor Address",
    "PO Number",
    "Notes",
]

# ── Sample pre-filled rows (to show non-empty cells) ─────────────────────────
SAMPLE_ROWS = [
    # Row 2 — completely empty (the "active row" for demo)
    {h: "" for h in HEADERS},

    # Row 3 — partially filled
    {
        "Vendor Name": "Acme Supplies Ltd",
        "Vendor GST Number": "",
        "Invoice Number": "INV-2024-001",
        "Invoice Date": "2024-01-15",
        "Due Date": "",
        "Item Description": "Office Furniture",
        "Quantity": "10",
        "Unit Price": "250.00",
        "Total Amount": "2500.00",
        "Currency": "USD",
        "Payment Terms": "Net 30",
        "Vendor Email": "",
        "Vendor Phone": "",
        "Vendor Address": "",
        "PO Number": "PO-9871",
        "Notes": "",
    },

    # Row 4 — fully filled example
    {
        "Vendor Name": "TechParts India Pvt Ltd",
        "Vendor GST Number": "27AADCT1234F1Z5",
        "Invoice Number": "INV-2024-042",
        "Invoice Date": "2024-02-10",
        "Due Date": "2024-03-10",
        "Item Description": "Electronic Components",
        "Quantity": "500",
        "Unit Price": "12.50",
        "Total Amount": "6250.00",
        "Currency": "INR",
        "Payment Terms": "Net 45",
        "Vendor Email": "billing@techparts.in",
        "Vendor Phone": "+91-22-40001234",
        "Vendor Address": "Plot 15, MIDC, Pune, Maharashtra 411019",
        "PO Number": "PO-3301",
        "Notes": "Urgent delivery required",
    },
]


def create_workbook(output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # ── Header row styling ─────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="center")

    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

    for row_idx, row_data in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, header in enumerate(HEADERS, start=1):
            val = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else None)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = 20

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = {
        "Vendor Name": 25,
        "Vendor GST Number": 22,
        "Invoice Number": 18,
        "Invoice Date": 14,
        "Due Date": 14,
        "Item Description": 28,
        "Quantity": 10,
        "Unit Price": 12,
        "Total Amount": 14,
        "Currency": 10,
        "Payment Terms": 16,
        "Vendor Email": 28,
        "Vendor Phone": 18,
        "Vendor Address": 35,
        "PO Number": 14,
        "Notes": 30,
    }

    for col_idx, header in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(header, 15)

    # ── Freeze header row ──────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Workbook saved: {output_path}")
    print(f"  Sheet: Invoices")
    print(f"  Columns ({len(HEADERS)}): {', '.join(HEADERS)}")
    print(f"  Rows: 1 header + {len(SAMPLE_ROWS)} data rows")
    print(f"  Row 2 is completely empty — use as active_row for testing.")


if __name__ == "__main__":
    output = Path(__file__).parent / "vendor_invoice.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    create_workbook(output)
